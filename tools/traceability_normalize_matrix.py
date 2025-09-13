import argparse
from pathlib import Path
import re
import shutil
import sys

from openpyxl import Workbook, load_workbook

CANONICAL_HEADER = [
    "Feature Name",
    "PRD Reference",
    "Arch Reference",
    "UI Reference",
    "ADR Reference",
    "Validation Method",
    "Tag",
]
LEGACY_HEADER = [
    "ID",
    "Module",
    "UI Element / Flow",
    "PRD Requirement (short)",
    "Architecture Mapping",
    "Interaction / Accessibility",
    "Telemetry / KPI",
    "Status",
    "Owner",
    "Notes",
    "UI Framework Mapping",
]
SEC_PAT = re.compile(r"§\s*\d+(?:\.\d+)*")


def extract_sec_tokens(text):
    if not text:
        return ""
    tokens = SEC_PAT.findall(str(text))
    tokens = [t.replace(" ", "") for t in tokens]
    # Deduplicate and join
    return "; ".join(sorted(set(tokens), key=tokens.index))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--trace", default="docs/traceability/Traceability_v4.1_prefilled.xlsx"
    )
    parser.add_argument("--sheet", default=None)
    parser.add_argument(
        "--out", default="docs/traceability/Traceability_v4.1_prefilled.xlsx"
    )
    parser.add_argument(
        "--backup", default="docs/traceability/Traceability_v4.1_prefilled.bak.xlsx"
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(args.trace)
    sheetnames = wb.sheetnames
    ws = None
    # Prefer Matrix, else Sheet1
    if args.sheet and args.sheet in sheetnames:
        ws = wb[args.sheet]
    elif "Matrix" in sheetnames:
        ws = wb["Matrix"]
    elif "Sheet1" in sheetnames:
        ws = wb["Sheet1"]
    else:
        print("No valid worksheet found.", file=sys.stderr)
        sys.exit(1)

    # Detect headers
    header_row = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    is_canonical = all(h in header_row for h in CANONICAL_HEADER)
    is_legacy = all(h in header_row for h in LEGACY_HEADER)
    # Find legacy header row index
    legacy_idx = None
    for i, cell in enumerate(ws[1]):
        if str(cell.value).strip() == "ID":
            legacy_idx = 1
            break
    # Find canonical header row index
    canon_idx = None
    for i, cell in enumerate(ws[1]):
        if str(cell.value).strip() == "Feature Name":
            canon_idx = 1
            break
    # If already canonical and no legacy, exit
    if is_canonical and not is_legacy:
        print("Already canonical. No legacy rows detected.")
        sys.exit(0)

    # Backup if needed
    if not args.dry_run and not Path(args.backup).exists():
        shutil.copy(args.trace, args.backup)

    # Collect legacy rows
    legacy_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        values = [str(cell.value).strip() if cell.value else "" for cell in row]
        if len(values) >= len(LEGACY_HEADER) and values[0] and values[1]:
            # Looks like legacy row
            legacy_rows.append(values)
        elif len(values) >= len(CANONICAL_HEADER) and values[0] and values[1]:
            # Looks like canonical row
            break

    # Convert legacy rows
    converted = []
    for vals in legacy_rows:
        # Map legacy columns
        module = vals[1]
        ui_elem = vals[2]
        prd_short = vals[3]
        arch_map = vals[4]
        ui_map = vals[10] if len(vals) > 10 else ""
        status = vals[7]
        feature_name = f"{module}: {ui_elem}".strip()
        prd_ref = extract_sec_tokens(prd_short)
        arch_ref = extract_sec_tokens(arch_map)
        ui_ref = extract_sec_tokens(ui_map)
        tag = status if status else "legacy-auto"
        converted.append(
            [
                feature_name,
                prd_ref,
                arch_ref,
                ui_ref,
                "",  # ADR Reference
                "",  # Validation Method
                tag,
            ]
        )

    # Collect canonical rows
    canonical_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        values = [str(cell.value).strip() if cell.value else "" for cell in row]
        if len(values) >= len(CANONICAL_HEADER) and values[0] and values[1]:
            canonical_rows.append(values)

    # Create new worksheet with canonical header
    new_wb = Workbook()
    # Remove default sheet if present
    if "Sheet" in new_wb.sheetnames:
        std = new_wb["Sheet"]
        new_wb.remove(std)
    new_ws = new_wb.create_sheet("Matrix")
    new_ws.append(CANONICAL_HEADER)
    for row in converted:
        new_ws.append(row)
    for row in canonical_rows:
        new_ws.append(row)

    if not args.dry_run:
        new_wb.save(args.out)
        print(f"Normalized matrix written to {args.out}")
    else:
        print("Dry run: no changes written.")

    # Re-run validator
    if not args.dry_run:
        import subprocess

        result = subprocess.run(
            [
                sys.executable,
                "tools/validate_traceability_md.py",
                "--prd",
                "docs/PRD/PRD_v4.0_unified_numbered.md",
                "--arch",
                "docs/Architecture/Architecture_v4.1_unified.md",
                "--ui",
                "docs/UIFramework/UIFramework_v4.0_unified_numbered.md",
                "--trace",
                args.out,
                "--out",
                "docs/traceability/Traceability_link_check.csv",
            ]
        )
        print("Validator run complete.")


if __name__ == "__main__":
    main()
