import argparse

from openpyxl import load_workbook


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--list", default="tools/legacy_waive_list.txt")
    parser.add_argument(
        "--trace", default="docs/traceability/Traceability_v4.1_prefilled.xlsx"
    )
    parser.add_argument("--sheet", default="Matrix")
    args = parser.parse_args()

    with open(args.list, encoding="utf-8") as f:
        waive_features = set(line.strip().lower() for line in f if line.strip())

    wb = load_workbook(args.trace)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [
        str(cell.value).strip().lower().replace(" ", "").replace("_", "")
        for cell in ws[1]
    ]
    feat_idx = None
    tag_idx = None
    for i, h in enumerate(headers):
        if "featurename" in h or "feature" in h:
            feat_idx = i
        if "tag" in h:
            tag_idx = i
    if tag_idx is None:
        ws.insert_cols(ws.max_column + 1)
        from openpyxl.cell.cell import MergedCell

        tag_header_cell = ws.cell(row=1, column=ws.max_column)
        if not isinstance(tag_header_cell, MergedCell):
            tag_header_cell.value = "Tag"
        tag_idx = ws.max_column - 1
    from openpyxl.cell.cell import MergedCell

    for row in ws.iter_rows(min_row=2):
        feat = str(row[feat_idx].value).strip().lower() if feat_idx is not None else ""
        tag_cell = row[tag_idx]
        if feat in waive_features and not isinstance(tag_cell, MergedCell):
            tag_val = str(tag_cell.value or "").strip()
            if "legacy-waived" not in tag_val:
                if tag_val:
                    tag_cell.value = tag_val + "; legacy-waived"
                else:
                    tag_cell.value = "legacy-waived"
    wb.save(args.trace)
    print("Legacy rows waived.")


if __name__ == "__main__":
    main()
