import logging
from typing import Optional
def _autodiscover_spec(path_hint: Optional[str], filename: str) -> Optional[str]:
    if path_hint and Path(path_hint).is_file():
        return str(Path(path_hint).resolve())
    for p in Path.cwd().rglob(filename):
        if p.is_file():
            return str(p.resolve())
    return None
import logging
from typing import Optional
import argparse
import json
import re
import difflib
import sys
import subprocess
from pathlib import Path
from openpyxl import load_workbook

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument('--prd', default='docs/PRD/PRD_v4.0_unified_numbered.md')
    p.add_argument('--arch', default='docs/Architecture/Architecture_v4.1_unified.md')
    p.add_argument('--ui', default='docs/UI_Framework/UIFramework_v4.0_unified_numbered.md')
    p.add_argument('--trace', default='docs/traceability/Traceability_v4.1_prefilled.xlsx')
    p.add_argument('--sheet', default='Matrix')
    p.add_argument('--out', default='tools/traceability_fix_report.json')
    p.add_argument('--apply', action='store_true')
    return p.parse_args()

def extract_section_keys(md_path):
    keys = {}
    pat = re.compile(r'^(#+)\s*(§?\s*[\d]+(?:\.[\d]+)*)(\s+.*)?$', re.MULTILINE)
    with open(md_path, encoding='utf-8') as f:
        for line in f:
            m = pat.match(line.strip())
            if m:
                raw = m.group(2)
                norm = '§' + re.sub(r'\s+', '', raw.replace('§',''))
                keys[norm] = line.strip()
    return keys

def normalize_token(t):
    t = t.strip().replace(' ', '')
    if not t.startswith('§'):
        t = '§' + t.lstrip('§')
    return t

def split_refs(cell):
    return [normalize_token(x) for x in re.split(r'[;,]', str(cell or '')) if x.strip()]

def find_closest(token, available):
    for i in range(len(token), 2, -1):
        prefix = token[:i]
        matches = [k for k in available if k.startswith(prefix)]
        if matches:
            return matches
    close = difflib.get_close_matches(token, available, n=3, cutoff=0.7)
    return close

def main():
    args = parse_args()
        prd_path = args.prd if Path(args.prd).is_file() else _autodiscover_spec(None, "PRD_v4.0_unified_numbered.md")
        arch_path = args.arch if Path(args.arch).is_file() else _autodiscover_spec(None, "Architecture_v4.1_unified.md")
        ui_path = args.ui if Path(args.ui).is_file() else _autodiscover_spec(None, "UIFramework_v4.0_unified_numbered.md")

        prd_keys = extract_section_keys(prd_path) if prd_path else {}
        arch_keys = extract_section_keys(arch_path) if arch_path else {}
        ui_keys = extract_section_keys(ui_path) if ui_path else None
        ui_skipped = ui_keys is None
        if ui_skipped:
            logging.warning("UIFramework spec not found; skipping UI refs (will not auto-fix UI mismatches).")
    wb = load_workbook(args.trace)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(cell.value).strip().lower().replace(' ','').replace('_','') for cell in ws[1]]
    colmap = {}
    for idx, h in enumerate(headers):
        if 'featurename' in h: colmap['Feature Name'] = idx
        if 'prdreference' in h: colmap['PRD Reference'] = idx
        if 'archreference' in h: colmap['Arch Reference'] = idx
        if 'uireference' in h: colmap['UI Reference'] = idx
        if 'adrreference' in h: colmap['ADR Reference'] = idx
        if 'validationmethod' in h: colmap['Validation Method'] = idx
    findings = []
    fixed = 0
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for colname, keys_map in [('PRD Reference', prd_keys), ('Arch Reference', arch_keys), ('UI Reference', ui_keys)]:
            colidx = colmap.get(colname)
            if colidx is None: continue
            cell = row[colidx].value
            tokens = split_refs(cell)
            new_tokens = list(tokens)
            for j, t in enumerate(tokens):
                if t not in keys_map:
                    candidates = find_closest(t, list(keys_map.keys()))
                    findings.append({'row': i, 'feature': row[colmap['Feature Name']].value, 'column': colname, 'bad_ref': t, 'candidates': candidates})
                    if args.apply and candidates:
                        chosen = None
                        for c in candidates:
                            if re.match(r'^§[\d]+(\.[\d]+)*$', c) and t.split('.')[0] == c.split('.')[0]:
                                chosen = c
                                break
                        if not chosen and candidates:
                            sim = difflib.SequenceMatcher(None, t, candidates[0]).ratio()
                            if sim >= 0.8:
                                chosen = candidates[0]
                        if chosen:
                            new_tokens[j] = chosen
                            fixed += 1
            if args.apply and new_tokens != tokens:
                row[colidx].value = '; '.join(new_tokens)
    if args.apply and fixed:
        wb.save(args.trace)
        print(f"Patched {fixed} refs in {args.trace}")
    report = {'missing_refs': findings, 'fixed': fixed, 'counts': {'PRD Reference':0,'Arch Reference':0,'UI Reference':0}, 'examples':[]}
    for f in findings:
        if f['column'] in report['counts']:
            report['counts'][f['column']] += 1
        report['examples'].append(f)
    with open(args.out, 'w') as f:
        json.dump(report, f, indent=2)
    print(f"Traceability ref scan complete. See {args.out}")
    if args.apply:
        print("Running traceability validator...")
        result = subprocess.run([
            sys.executable, "tools/validate_traceability_md.py",
            "--prd", args.prd,
            "--arch", args.arch,
            "--ui", args.ui,
            "--trace", args.trace,
            "--out", "docs/traceability/Traceability_link_check.csv"
        ])
        if result.returncode == 0:
            print("Traceability validation PASSED.")
        else:
            print("Traceability validation FAILED.")
        sys.exit(result.returncode)

if __name__ == "__main__":
    main()
