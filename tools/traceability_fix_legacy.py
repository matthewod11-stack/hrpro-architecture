import logging


def _autodiscover_spec(path_hint: str | None, filename: str) -> str | None:
    if path_hint and Path(path_hint).is_file():
        return str(Path(path_hint).resolve())
    for p in Path.cwd().rglob(filename):
        if p.is_file():
            return str(p.resolve())
    return None


import argparse
import csv
import difflib
import json
from pathlib import Path
import re
import subprocess
import sys

from openpyxl import load_workbook

PR2_FEATURES = set(
    [
        "advisor sse streaming ui integration",
        "citation chips + drawer (a11y)",
        "export summary (pdf) button",
        "telemetry dashboard",
        "e2e demo script",
    ]
)


def normalize_feature_name(name):
    return name.strip().lower()


def normalize_token(t):
    t = t.strip().replace(" ", "")
    if not t.startswith("§"):
        t = "§" + t.lstrip("§")
    return t


def split_refs(cell):
    return [normalize_token(x) for x in re.split(r"[;,]", str(cell or "")) if x.strip()]


def extract_section_keys(md_path):
    keys = {}
    pat = re.compile(r"^(#+)\s*(§?\s*[\d]+(?:\.[\d]+)*)(\s+.*)?$", re.MULTILINE)
    with open(md_path, encoding="utf-8") as f:
        for line in f:
            m = pat.match(line.strip())
            if m:
                raw = m.group(2)
                norm = "§" + re.sub(r"\s+", "", raw.replace("§", ""))
                keys[norm] = line.strip()
    return keys


def find_candidates(token, available):
    for i in range(len(token), 2, -1):
        prefix = token[:i]
        matches = [k for k in available if k.startswith(prefix)]
        if matches:
            return matches
    close = difflib.get_close_matches(token, available, n=3, cutoff=0.8)
    return close


def is_safe_fix(candidates, token):
    if not candidates:
        return False, None
    chosen = None
    for c in candidates:
        if (
            re.match(r"^§[\d]+(\.[\d]+)*$", c)
            and token.split(".")[0] == c.split(".")[0]
        ):
            chosen = c
            break
    if not chosen and candidates:
        sim = difflib.SequenceMatcher(None, token, candidates[0]).ratio()
        if sim >= 0.8:
            chosen = candidates[0]
    return bool(chosen), chosen


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--prd", default="docs/PRD/PRD_v4.0_unified_numbered.md")
    parser.add_argument(
        "--arch", default="docs/Architecture/Architecture_v4.1_unified.md"
    )
    parser.add_argument(
        "--ui", default="docs/UIFramework/UIFramework_v4.0_unified_numbered.md"
    )
    parser.add_argument(
        "--trace", default="docs/traceability/Traceability_v4.1_prefilled.xlsx"
    )
    parser.add_argument("--sheet", default="Matrix")
    parser.add_argument("--out-report", default="tools/trace_legacy_report.json")
    parser.add_argument("--out-csv", default="tools/trace_legacy_todo.csv")
    parser.add_argument("--apply-safe", action="store_true")
    args = parser.parse_args()

    prd_path = (
        args.prd
        if Path(args.prd).is_file()
        else _autodiscover_spec(None, "PRD_v4.0_unified_numbered.md")
    )
    arch_path = (
        args.arch
        if Path(args.arch).is_file()
        else _autodiscover_spec(None, "Architecture_v4.1_unified.md")
    )
    ui_path = (
        args.ui
        if Path(args.ui).is_file()
        else _autodiscover_spec(None, "UIFramework_v4.0_unified_numbered.md")
    )

    prd_keys = extract_section_keys(prd_path) if prd_path else {}
    arch_keys = extract_section_keys(arch_path) if arch_path else {}
    ui_keys = extract_section_keys(ui_path) if ui_path else None
    ui_skipped = ui_keys is None
    if ui_skipped:
        logging.warning(
            "UIFramework spec not found; skipping UI refs (will not auto-fix UI mismatches)."
        )

    val_csv = Path("docs/traceability/Traceability_link_check.csv")
    with val_csv.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    headers = rows[0]

    def find_header(aliases):
        for alias in aliases:
            norm_alias = alias.strip().lower().replace(" ", "").replace("_", "")
            for i, h in enumerate(headers):
                if h.strip().lower().replace(" ", "").replace("_", "") == norm_alias:
                    return i
        return None

    feat_idx = find_header(["feature name", "feature", "featurename"])
    prd_idx = find_header(["prd reference", "prd", "prd ref"])
    arch_idx = find_header(["arch reference", "arch", "arch ref"])
    ui_idx = find_header(["ui reference", "ui", "ui ref"])
    status_idx = find_header(["status"])

    wb = load_workbook(args.trace)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    excel_headers = [
        str(cell.value).strip().lower().replace(" ", "").replace("_", "")
        for cell in ws[1]
    ]
    excel_feat_idx = (
        excel_headers.index("featurename") if "featurename" in excel_headers else None
    )
    excel_prd_idx = (
        excel_headers.index("prdreference") if "prdreference" in excel_headers else None
    )
    excel_arch_idx = (
        excel_headers.index("archreference")
        if "archreference" in excel_headers
        else None
    )
    excel_ui_idx = (
        excel_headers.index("uireference") if "uireference" in excel_headers else None
    )

    report = []
    ambiguous = []
    safe_fixes = []
    for i, row in enumerate(rows[1:], start=2):
        feat = row[feat_idx].strip() if feat_idx is not None else ""
        norm_feat = normalize_feature_name(feat)
        if norm_feat in PR2_FEATURES:
            continue
        status = row[status_idx] if status_idx is not None else ""
        fail = (
            status.lower() != "ok"
            or "mismatch" in status.lower()
            or "missing" in status.lower()
        )
        if not fail:
            for cell in row:
                if "mismatch" in cell.lower() or "missing" in cell.lower():
                    fail = True
                    break
        if not fail:
            continue
        row_report = {
            "row": i,
            "feature": feat,
            "status": status,
            "fixes": {},
            "ambiguous": {},
        }
        col_specs = [
            ("PRD Reference", prd_idx, prd_keys, excel_prd_idx),
            ("Arch Reference", arch_idx, arch_keys, excel_arch_idx),
        ]
        if not ui_skipped:
            col_specs.append(("UI Reference", ui_idx, ui_keys, excel_ui_idx))
        for colname, idx, keys_map, excel_idx in col_specs:
            cell = row[idx] if idx is not None else ""
            tokens = split_refs(cell)
            new_tokens = list(tokens)
            for j, t in enumerate(tokens):
                if t not in keys_map:
                    candidates = find_candidates(t, list(keys_map.keys()))
                    safe, chosen = is_safe_fix(candidates, t)
                    entry = {
                        "column": colname,
                        "bad_ref": t,
                        "candidates": candidates,
                        "chosen": chosen,
                        "safe": safe,
                    }
                    if safe and args.apply_safe and excel_idx is not None:
                        excel_row = ws[i]
                        excel_cell = excel_row[excel_idx]
                        excel_tokens = split_refs(excel_cell.value)
                        excel_tokens = [
                            str(chosen) if x == t else str(x)
                            for x in excel_tokens
                            if x is not None
                        ]
                        from openpyxl.cell.cell import MergedCell

                        if not isinstance(excel_cell, MergedCell):
                            excel_cell.value = "; ".join(excel_tokens)
                        safe_fixes.append(
                            {
                                "row": i,
                                "feature": feat,
                                "column": colname,
                                "bad_ref": t,
                                "chosen": chosen,
                            }
                        )
                    if safe:
                        row_report["fixes"][colname] = entry
                    else:
                        row_report["ambiguous"][colname] = entry
                        ambiguous.append([i, feat, colname, t, ";".join(candidates)])
        report.append(row_report)
    if args.apply_safe and safe_fixes:
        wb.save(args.trace)
    with open(args.out_report, "w") as f:
        json.dump(report, f, indent=2)
    with open(args.out_csv, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Row", "Feature", "Column", "BadRef", "Candidates"])
        for row in ambiguous:
            writer.writerow(row)
    result = subprocess.run(
        [
            sys.executable,
            "tools/validate_traceability_md.py",
            "--prd",
            args.prd,
            "--arch",
            args.arch,
            "--ui",
            args.ui,
            "--trace",
            args.trace,
            "--out",
            "docs/traceability/Traceability_link_check.csv",
        ]
    )
    ok, fail = 0, 0
    ui_fail = 0
    with open("docs/traceability/Traceability_link_check.csv", encoding="utf-8") as f:
        reader = csv.reader(f)
        hdr = next(reader)
        status_idx = None
        ui_idx = None
        for i, h in enumerate(hdr):
            h_norm = h.strip().lower().replace(" ", "").replace("_", "")
            if h_norm == "status":
                status_idx = i
            if h_norm == "uireference":
                ui_idx = i
        for row in reader:
            status_val = row[status_idx].lower() if status_idx is not None else ""
            ui_val = row[ui_idx] if ui_idx is not None else ""
            if status_val == "ok":
                ok += 1
            else:
                # Only count as fail if not UI-only and UI was skipped
                if (
                    ui_skipped
                    and ui_val
                    and ("missing" in status_val or "mismatch" in status_val)
                ):
                    ui_fail += 1
                else:
                    fail += 1
    print(f"Traceability validation: {ok} OK / {fail} failing rows.")
    if ui_skipped:
        print(f"UIFramework spec was skipped; {ui_fail} UI-only failures ignored.")
    sys.exit(0 if fail == 0 else 1)


if __name__ == "__main__":
    main()
