from pathlib import Path
import subprocess
import sys

from openpyxl import load_workbook

PR2_ROWS = [
    {
        "Feature Name": "Advisor SSE streaming UI integration",
        "PRD Reference": "§1.4.5",
        "Arch Reference": "§4.1; §4.2; §11.1",
        "UI Reference": "§7.6",
        "ADR Reference": "ADR-0001; ADR-0003",
        "Validation Method": "pytest tests/test_advisor_endpoint.py; manual demo",
    },
    {
        "Feature Name": "Citation chips + drawer (a11y)",
        "PRD Reference": "§1.5.13",
        "Arch Reference": "§4.1; §4.2",
        "UI Reference": "§7.6",
        "ADR Reference": "ADR-0004",
        "Validation Method": "manual Streamlit demo; tests/test_advisor_drawer.py",
    },
    {
        "Feature Name": "Export Summary (PDF) button",
        "PRD Reference": "§1.4.5",
        "Arch Reference": "§11.1",
        "UI Reference": "§7.6",
        "ADR Reference": "ADR-0001",
        "Validation Method": "tools/smoke_e2e.py; manual demo",
    },
    {
        "Feature Name": "Telemetry dashboard",
        "PRD Reference": "§1.4.5",
        "Arch Reference": "§4.1; §11.1",
        "UI Reference": "§7.6",
        "ADR Reference": "ADR-0003",
        "Validation Method": "manual Streamlit demo",
    },
    {
        "Feature Name": "E2E demo script",
        "PRD Reference": "§1.4.5",
        "Arch Reference": "§11.1",
        "UI Reference": "§7.6",
        "ADR Reference": "ADR-0003",
        "Validation Method": "tools/smoke_e2e.py",
    },
]

TRACE_PATH = Path("docs/traceability/Traceability_v4.1_prefilled.xlsx")
VALIDATOR_CMD = [
    sys.executable,
    "tools/validate_traceability_md.py",
    "--prd",
    "docs/PRD/PRD_v4.0_unified_numbered.md",
    "--arch",
    "docs/Architecture/Architecture_v4.1_unified.md",
    "--ui",
    "docs/UIFramework/UIFramework_v4.0_unified_numbered.md",
    "--trace",
    str(TRACE_PATH),
    "--out",
    "docs/traceability/Traceability_link_check.csv",
]

REQUIRED_HEADERS = [
    "Feature Name",
    "PRD Reference",
    "Arch Reference",
    "UI Reference",
    "ADR Reference",
    "Validation Method",
]


def _normalize_header(h):
    return h.strip().lower().replace(" ", "").replace("_", "")


def _find_or_create_headers(ws, headers):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        norm = [_normalize_header(str(cell or "")) for cell in row]
        if all(_normalize_header(h) in norm for h in headers):
            return {h: norm.index(_normalize_header(h)) for h in headers}, i + 1
    ws.insert_rows(1)
    for j, h in enumerate(headers):
        ws.cell(row=1, column=j + 1, value=h)
    return {h: j for j, h in enumerate(headers)}, 1


def _existing_feature_names(ws, col_idx, header_row):
    names = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        val = row[col_idx] if col_idx < len(row) else None
        if val:
            names.add(str(val))
    return names


def _append_row(ws, row_dict, colmap, header_row):
    row = [row_dict.get(h, "") for h in REQUIRED_HEADERS]
    ws.append(row)


def main():
    wb = load_workbook(TRACE_PATH)
    ws = wb["Matrix"] if "Matrix" in wb.sheetnames else wb[wb.sheetnames[0]]
    colmap, header_row = _find_or_create_headers(ws, REQUIRED_HEADERS)
    feat_col = colmap["Feature Name"]
    existing = _existing_feature_names(ws, feat_col, header_row)
    added = 0
    for row in PR2_ROWS:
        if row["Feature Name"] not in existing:
            _append_row(ws, row, colmap, header_row)
            added += 1
    if added:
        wb.save(TRACE_PATH)
        print(f"Appended {added} PR2 rows to {TRACE_PATH}")
    else:
        print("No new PR2 rows needed (already present).")
    if "--validate" in sys.argv:
        print("Running traceability validator...")
        result = subprocess.run(VALIDATOR_CMD)
        if result.returncode == 0:
            print("Traceability validation PASSED.")
        else:
            print("Traceability validation FAILED.")
        sys.exit(result.returncode)


if __name__ == "__main__":
    main()
