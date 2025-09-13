import argparse
from collections import defaultdict
import csv
import difflib
import json
import re
import sys

from openpyxl import load_workbook

PR2_FEATURES = {
    "advisor sse streaming ui integration",
    "citation chips + drawer (a11y)",
    "export summary (pdf) button",
    "telemetry dashboard",
    "e2e demo script",
}

KEYWORD_BOOST = {
    "landing": 1.5,
    "home": 1.2,
    "module": 1.1,
    "advisor": 1.5,
    "typing": 1.2,
    "citations": 1.2,
    "prompt": 1.2,
    "streaming": 1.2,
    "dashboard": 1.5,
    "grid": 1.2,
    "tiles": 1.1,
    "analytics": 1.5,
    "demographic": 1.5,
    "breakdown": 1.2,
    "filter": 1.1,
    "accessibility": 1.5,
    "keyboard": 1.3,
    "contrast": 1.3,
    "aria": 1.3,
    "export": 1.2,
}

SECTION_PAT = re.compile(r"^(#+)\s*(§?\s*[\d]+(?:\.[\d]+)*)(\s+.*)?$", re.MULTILINE)


def normalize_key(raw):
    raw = raw.strip().replace(" ", "")
    if not raw.startswith("§"):
        raw = "§" + raw.lstrip("§")
    return raw


def parse_spec(md_path):
    keys = {}
    keyword_index = defaultdict(set)
    with open(md_path, encoding="utf-8") as f:
        for line in f:
            m = SECTION_PAT.match(line.strip())
            if m:
                key = normalize_key(m.group(2))
                heading = line.strip()
                keys[key] = heading
                tokens = [t.lower() for t in re.split(r"\W+", heading) if len(t) >= 3]
                for t in tokens:
                    keyword_index[t].add(key)
    return keys, keyword_index


def get_feature_tokens(feature):
    tokens = [t.lower() for t in re.split(r"\W+", feature) if len(t) >= 3]
    return tokens


def rank_candidates(feature, candidates, spec_keys, keyword_index):
    tokens = get_feature_tokens(feature)
    scores = {}
    for cand in candidates:
        score = 0.0
        heading = spec_keys.get(cand, "")
        heading_tokens = [t.lower() for t in re.split(r"\W+", heading) if len(t) >= 3]
        overlap = len(set(tokens) & set(heading_tokens))
        score += overlap
        for t in tokens:
            if t in heading_tokens:
                score += 0.1
            score += KEYWORD_BOOST.get(t, 0)
        # Similarity boost
        score += difflib.SequenceMatcher(None, cand, feature).ratio()
        scores[cand] = score
    sorted_cands = sorted(scores.items(), key=lambda x: -x[1])
    return sorted_cands


def find_best_key(feature, bad_ref, candidates, spec_keys, keyword_index):
    # Direct candidate match
    valid_cands = [c for c in candidates if c in spec_keys]
    if not valid_cands:
        # Fuzzy match against all spec keys
        valid_cands = difflib.get_close_matches(
            bad_ref, list(spec_keys.keys()), n=5, cutoff=0.8
        )
    if not valid_cands:
        # Keyword overlap
        tokens = get_feature_tokens(feature)
        key_scores = defaultdict(float)
        for t in tokens:
            for k in keyword_index.get(t, []):
                key_scores[k] += 1 + KEYWORD_BOOST.get(t, 0)
        if key_scores:
            valid_cands = [
                k for k, _ in sorted(key_scores.items(), key=lambda x: -x[1])[:3]
            ]
    if not valid_cands:
        # Numeric fallback
        m = re.match(r"^§([\d]+(\.[\d]+)*)", bad_ref)
        if m:
            parts = m.group(1).split(".")
            for i in range(len(parts) - 1, 0, -1):
                prefix = "§" + ".".join(parts[:i])
                if prefix in spec_keys:
                    valid_cands = [prefix]
                    break
    if not valid_cands:
        return None, "no valid candidates", 0.0
    # Rank
    ranked = rank_candidates(feature, valid_cands, spec_keys, keyword_index)
    if not ranked:
        return None, "no ranked candidates", 0.0
    top, top_score = ranked[0]
    if len(ranked) == 1 or (len(ranked) > 1 and top_score - ranked[1][1] >= 0.15):
        return top, "clear lead", top_score
    return None, "ambiguous", top_score


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--trace", default="docs/traceability/Traceability_v4.1_prefilled.xlsx"
    )
    parser.add_argument("--sheet", default="Matrix")
    parser.add_argument("--prd", default="docs/PRD/PRD_v4.0_unified_numbered.md")
    parser.add_argument(
        "--arch", default="docs/Architecture/Architecture_v4.1_unified.md"
    )
    parser.add_argument(
        "--ui", default="docs/UIFramework/UIFramework_v4.0_unified_numbered.md"
    )
    parser.add_argument("--todo", default="tools/trace_legacy_todo.csv")
    parser.add_argument("--max-edits", type=int, default=999)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--out-report", default="tools/trace_legacy_autopatch.json")
    args = parser.parse_args()

    prd_keys, prd_kw = parse_spec(args.prd)
    arch_keys, arch_kw = parse_spec(args.arch)
    ui_keys, ui_kw = parse_spec(args.ui)
    spec_map = {
        "PRD Reference": (prd_keys, prd_kw),
        "Arch Reference": (arch_keys, arch_kw),
        "UI Reference": (ui_keys, ui_kw),
    }

    with open(args.todo, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        todo_rows = list(reader)

    wb = load_workbook(args.trace)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [
        str(cell.value).strip().lower().replace(" ", "").replace("_", "")
        for cell in ws[1]
    ]
    colmap = {}
    for idx, h in enumerate(headers):
        if "featurename" in h:
            colmap["Feature Name"] = idx
        if "prdreference" in h:
            colmap["PRD Reference"] = idx
        if "archreference" in h:
            colmap["Arch Reference"] = idx
        if "uireference" in h:
            colmap["UI Reference"] = idx
    patched = []
    ambiguous = []
    edits = 0
    for todo_row in todo_rows:
        if edits >= args.max_edits:
            break
        feat = todo_row.get("Feature", "").strip()
        norm_feat = feat.lower()
        if norm_feat in PR2_FEATURES:
            continue
        col = todo_row.get("Column", "").strip()
        bad_ref = todo_row.get("BadRef", "").strip()
        candidates = [
            c.strip() for c in todo_row.get("Candidates", "").split(";") if c.strip()
        ]
        spec_keys, keyword_index = spec_map.get(col, ({}, {}))
        chosen, reason, score = find_best_key(
            feat, bad_ref, candidates, spec_keys, keyword_index
        )
        try:
            excel_row_idx = int(todo_row.get("Row", "0"))
            excel_col_idx = colmap.get(col)
            if excel_col_idx is None:
                ambiguous.append(
                    {
                        "row": excel_row_idx,
                        "column": col,
                        "feature": feat,
                        "bad_ref": bad_ref,
                        "candidates": candidates,
                        "reason": "column not found",
                    }
                )
                continue
            excel_row = list(
                ws.iter_rows(min_row=excel_row_idx, max_row=excel_row_idx)
            )[0]
            excel_cell = excel_row[excel_col_idx]
            before = str(excel_cell.value)
            after = before
            if chosen:
                tokens = [t.strip() for t in re.split(r"[;,]", before) if t.strip()]
                if args.force:
                    after = chosen
                else:
                    after_tokens = [chosen if t == bad_ref else t for t in tokens]
                    after = "; ".join(after_tokens)
                if after != before:
                    edits += 1
                    if not args.dry_run:
                        excel_cell.value = after
                patched.append(
                    {
                        "row": excel_row_idx,
                        "column": col,
                        "feature": feat,
                        "before": before,
                        "after": after,
                        "chosen_key": chosen,
                        "reason": reason,
                        "score": score,
                    }
                )
            else:
                ambiguous.append(
                    {
                        "row": excel_row_idx,
                        "column": col,
                        "feature": feat,
                        "bad_ref": bad_ref,
                        "candidates": candidates,
                        "reason": reason,
                    }
                )
        except Exception as e:
            ambiguous.append(
                {
                    "row": todo_row.get("Row", ""),
                    "column": col,
                    "feature": feat,
                    "bad_ref": bad_ref,
                    "candidates": candidates,
                    "reason": f"exception: {e}",
                }
            )
    if not args.dry_run and edits:
        wb.save(args.trace)
    with open(args.out_report, "w") as f:
        json.dump(
            {"patched": patched, "ambiguous": ambiguous, "edits": edits}, f, indent=2
        )
    # Re-run validator
    result = None
    if not args.dry_run:
        import subprocess

        result = subprocess.run(
            [
                sys.executable,
                "tools/validate_traceability_md.py",
                "--prd",
                args.prd,
                "--arch",
                args.arch,
                "--ui",
                args.ui,
                "--trace",
                args.trace,
                "--out",
                "docs/traceability/Traceability_link_check.csv",
            ]
        )
    # Print summary
    ok, fail = 0, 0
    with open("docs/traceability/Traceability_link_check.csv", encoding="utf-8") as f:
        reader = csv.reader(f)
        hdr = next(reader)
        status_idx = None
        for i, h in enumerate(hdr):
            if h.strip().lower().replace(" ", "").replace("_", "") == "status":
                status_idx = i
                break
        for row in reader:
            if status_idx is not None and row[status_idx].lower() == "ok":
                ok += 1
            else:
                fail += 1
    print(f"Validator: {ok} OK / {fail} failing")
    sys.exit(0 if fail == 0 else 1)


if __name__ == "__main__":
    main()
