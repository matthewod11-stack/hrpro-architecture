import argparse
import json
import re
import sys

from openpyxl import load_workbook


def normalize_col(col):
    c = col.strip().lower().replace(" ", "").replace("_", "")
    if "prd" in c:
        return "prdreference"
    if "arch" in c:
        return "archreference"
    if "ui" in c:
        return "uireference"
    if "feature" in c:
        return "featurename"
    return c


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--trace", default="docs/traceability/Traceability_v4.1_prefilled.xlsx"
    )
    parser.add_argument("--sheet", default="Matrix")
    parser.add_argument("--ambig", default="tools/trace_legacy_autopatch.json")
    parser.add_argument("--map", default="tools/trace_force_map.json")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--out", default="tools/trace_force_patch_report.json")
    args = parser.parse_args()

    with open(args.ambig, encoding="utf-8") as f:
        ambig = json.load(f).get("ambiguous", [])
    with open(args.map, encoding="utf-8") as f:
        force_map = json.load(f).get("entries", [])

    # Build lookup of ambiguous cells
    ambig_lookup = {}
    for a in ambig:
        row = int(a["row"])
        col = normalize_col(a["column"])
        ambig_lookup[(row, col)] = a

    wb = load_workbook(args.trace)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [
        str(cell.value).strip().lower().replace(" ", "").replace("_", "")
        for cell in ws[1]
    ]
    colmap = {h: i for i, h in enumerate(headers)}

    patched = []
    skipped = []
    warnings = []
    for entry in force_map:
        row = int(entry["row"])
        col = normalize_col(entry["column"])
        force_ref = entry["force_ref"].strip()
        if (row, col) not in ambig_lookup:
            warnings.append(
                f"Skipping row {row}, column {entry['column']}: not found in ambiguous list."
            )
            skipped.append(
                {"row": row, "column": entry["column"], "reason": "not in ambiguous"}
            )
            continue
        excel_col_idx = colmap.get(col)
        if excel_col_idx is None:
            warnings.append(f"Column {entry['column']} not found in sheet headers.")
            skipped.append(
                {"row": row, "column": entry["column"], "reason": "header not found"}
            )
            continue
        try:
            excel_row = list(ws.iter_rows(min_row=row, max_row=row))[0]
            excel_cell = excel_row[excel_col_idx]
            before = str(excel_cell.value)
            tokens = [t.strip() for t in re.split(r"[;,]", before) if t.strip()]
            # Idempotency: if already present, do not change
            if force_ref in tokens and len(tokens) == 1:
                after = before
            else:
                # Optionally preserve other valid tokens
                new_tokens = [force_ref] + [t for t in tokens if t != force_ref]
                after = "; ".join(sorted(set(new_tokens), key=new_tokens.index))
                if not args.dry_run:
                    from openpyxl.cell.cell import MergedCell

                    if not isinstance(excel_cell, MergedCell):
                        excel_cell.value = after
            patched.append(
                {
                    "row": row,
                    "column": entry["column"],
                    "before": before,
                    "after": after,
                }
            )
        except Exception as e:
            warnings.append(
                f"Exception patching row {row}, column {entry['column']}: {e}"
            )
            skipped.append({"row": row, "column": entry["column"], "reason": str(e)})
    if not args.dry_run and patched:
        wb.save(args.trace)
    with open(args.out, "w") as f:
        json.dump(
            {"patched": patched, "skipped": skipped, "warnings": warnings}, f, indent=2
        )
    # Re-run validator
    result = None
    if not args.dry_run:
        import subprocess

        result = subprocess.run(
            [
                sys.executable,
                "tools/validate_traceability_md.py",
                "--prd",
                "docs/PRD/PRD_v4.0_unified_numbered.md",
                "--arch",
                "docs/Architecture/Architecture_v4.1_unified.md",
                "--ui",
                "docs/UIFramework/UIFramework_v4.0_unified_numbered.md",
                "--trace",
                args.trace,
                "--out",
                "docs/traceability/Traceability_link_check.csv",
            ]
        )
    # Print summary
    ok, fail = 0, 0
    with open("docs/traceability/Traceability_link_check.csv", encoding="utf-8") as f:
        reader = iter(f)
        hdr = next(reader)
        status_idx = None
        hdrs = [
            h.strip().lower().replace(" ", "").replace("_", "") for h in hdr.split(",")
        ]
        for i, h in enumerate(hdrs):
            if h == "status":
                status_idx = i
                break
        for line in reader:
            row = line.strip().split(",")
            if (
                status_idx is not None
                and len(row) > status_idx
                and row[status_idx].lower() == "ok"
            ):
                ok += 1
            else:
                fail += 1
    print(f"Validator: {ok} OK / {fail} failing")
    sys.exit(0 if fail == 0 else 1)


if __name__ == "__main__":
    main()
